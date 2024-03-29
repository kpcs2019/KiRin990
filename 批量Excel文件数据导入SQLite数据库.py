from random import choice, randrange
from string import digits, ascii_letters
from os import listdir,mkdir
from os.path import isdir
import sqlite3
from time import time
from openpyxl import Workbook, load_workbook


def generateRandomData():
	if not isdir("xlsxs"):
		mkdir("xlsxs")
    #total表示记录总条数
	global total
	characters = digits+ascii_letters
	for i in range(5):
		xlsName = 'xlsxs\\'+str(i)+'.xlsx'
        #随机数，每个xlsx文件的行数不一样
		totalLines = randrange(10**2)
		wb = Workbook()
		ws = wb.worksheets[0]
        #表头
		ws.append(['字段一', '字段二', '字段三', '字段四', '字段五'])
        #随机数据，每行5个字段，每个字段30个字符
		for j in range(totalLines):
			line = [''.join((choice(characters) for ii in range(10))) for jj in range(5)]
			ws.append(line)
			total += 1
        #保存xlsx文件
		wb.save(xlsName)


#针对每个xlsx文件的生成器
def eachXlsx(xlsxFn):
    wb = load_workbook(xlsxFn)
    ws = wb.worksheets[0]
    for index, row in enumerate(ws.rows):
        #忽略表头
        if index == 0:
            continue
        yield tuple(map(lambda x:x.value, row))


#导入
def xlsx2sqlite():
    #获取所有xlsx文件
    xlsxs = ('xlsxs\\'+fn for fn in listdir('xlsxs'))
    #连接数据库，创建游标
    conn = sqlite3.connect('data.db')
    cur = conn.cursor()
    for xlsx in xlsxs:
        #批量导入，减少提交事务的次数，可以提高速度
        sql = 'insert into fromxlsx values(?,?,?,?,?)'
        cur.executemany(sql, eachXlsx(xlsx))
        conn.commit()


total = 0


generateRandomData()

start = time()
xlsx2sqlite()
delta = time()-start
print('导入用时：', delta)
print('导入速度（条/秒）：', total/delta)

# conn = sqlite3.connect(":memory:")
# cur = conn.cursor()
# cur.execute("CREATE TABLE people (name,age)")
# who = [("董付国",38),("Jack",51),("刘清涛",23)]

# cur.executemany("INSERT INTO people VALUES (?,?)",(who))
# cur.execute("SELECT * FROM people WHERE name='刘清涛' AND age=23")
# print(cur.fetchone())

# cur.execute("SELECT * FROM people")
# # cur.execute("SELECT * FROM people WHERE name=:who AND age=:age",{"who":who,"age":age})
# print(cur.fetchmany(2))
